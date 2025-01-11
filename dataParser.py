import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook
class dataParser:
    def __init__(self, fileName , rowTypes):
        self.fileName = fileName
        self.rowTypes = rowTypes
        self.wb = load_workbook(self.fileName)
        self.ws = self.wb.worksheets[0]
        self.colNames = []
        for i in range(1 , len(rowTypes)+1):
            self.colNames.append(self.ws.cell(row = 1 , column = i).value)
        self.dataList = []
        self.createList()
    def createRow(self , row_id):
        tmp_row = []
        for  i in range(1,len(self.rowTypes) + 1):
            tmp_row.append(self.ws.cell(row=row_id, column=i).value)
        return tmp_row
    def createList(self):
        for i in range(2, self.ws.max_row + 1):
            self.dataList.append(self.createRow(i))

